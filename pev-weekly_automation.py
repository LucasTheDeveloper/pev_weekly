from openpyxl import load_workbook
import xlsxwriter
import re
from openpyxl.drawing.image import Image
from datetime import datetime, time, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles import Border
import tabula
import pandas as pd
import time as t



def create_new_sheet(wb):
    new_sheet_name = "created_sheet_" + datetime.now().strftime("%Y%m%d_%H%M%S")
    ws = wb.create_sheet(new_sheet_name)
    ws.sheet_view.zoomScale = 70
    #get the previously added sheet (assuming it's the last sheet)
    previous_sheet_name = wb.sheetnames[-2] if len(wb.sheetnames) > 1 else None
    previous_sheet = wb[previous_sheet_name]

    if previous_sheet:
    # Copy data from the previous sheet to the new sheet, including style
        for row in previous_sheet.iter_rows(min_row=1, max_row=previous_sheet.max_row, min_col=1, max_col=previous_sheet.max_column):
         for cell in row:
             new_cell = ws[cell.coordinate]
             new_cell.value = cell.value

            # Copy cell formatting from previous sheet
             new_cell.font = Font(
                name=cell.font.name,
                size=cell.font.size,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color,
                underline=cell.font.underline,
                strikethrough=cell.font.strikethrough,
                vertAlign=cell.font.vertAlign,
            )
                # Set text wrapping to fit text within the cell
             new_cell.alignment = Alignment(wrapText=True)

             if cell.fill is not None:
                    new_cell.fill = PatternFill(start_color=cell.fill.start_color, end_color=cell.fill.end_color, fill_type=cell.fill.fill_type)
             # Copy cell borders
        # Copy cell borders
             if cell.border is not None:
                 new_cell.border = Border(
                     left=cell.border.left,
                     right=cell.border.right,
                     top=cell.border.top,
                     bottom=cell.border.bottom,
                    )
             ws.row_dimensions[cell.row].height = previous_sheet.row_dimensions[cell.row].height   
             ws.column_dimensions[cell.column_letter].width = previous_sheet.column_dimensions[cell.column_letter].width


    #hide column D, I and N
    ws.column_dimensions['D'].hidden = True
    ws.column_dimensions['I'].hidden = True
    ws.column_dimensions['N'].hidden = True
    
    # Clearing the data in column C
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=3, max_col=3):
        for cell in row:
            cell.value = None
    
    # Clearing the data in column F
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=6, max_col=6):
        for cell in row:
            cell.value = None

    # Clearing the data in column G
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=7, max_col=7):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column J
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=10, max_col=10):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column L
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=12, max_col=12):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column M
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=13, max_col=13):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column O
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=15, max_col=15):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column P
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=16, max_col=16):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column V
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=22, max_col=22):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column Y
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=25, max_col=25):
        for cell in row:
            cell.value = None
    
     # Clearing the data in column AB
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=28, max_col=28):
        for cell in row:
            cell.value = None


    #copy data from B5 to B15 and paste it into C5 to C15
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=15, min_col=2, max_col=2), start=5):
        ws.cell(row=i, column=3).value = row[0].value
    
     # Clearing the data in column B
    for row in ws.iter_rows(min_row=5, max_row=15, min_col=2, max_col=2):
        for cell in row:
            cell.value = None
    #copy data from auto.xlsx to the current sheet
    # Open the auto.xlsx file and get the value from cell D62
    auto_wb = load_workbook('auto.xlsx')
    auto_sheet = auto_wb.active  # data is in the active sheet of auto.xlsx
    

    #copy for Hartenbros
    auto_value_D13 = auto_sheet['D13'].value
    auto_value_A25 = auto_sheet['A25'].value
    integer_part_A25 = int(re.search(r'\d+', auto_value_A25).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_B5 = auto_sheet['B5'].value
    auto_value_D35 = auto_sheet['D35'].value
    auto_value_D34 = auto_sheet['D34'].value
   
    #copy for Eagles Landing
    auto_value_G63 = auto_sheet['G63'].value
    auto_value_A74 = auto_sheet['A74'].value
    integer_part_A74 = int(re.search(r'\d+', auto_value_A74).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_G54= auto_sheet['G54'].value
    auto_value_D84 = auto_sheet['D84'].value
    auto_value_D83 = auto_sheet['D83'].value

    #Copy for Irene
    auto_value_D111 = auto_sheet['D111'].value
    auto_value_A123 = auto_sheet['A123'].value
    integer_part_A123 = int(re.search(r'\d+', auto_value_A123).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_I103= auto_sheet['I103'].value
    auto_value_D133 = auto_sheet['D133'].value
    auto_value_D132 = auto_sheet['D132'].value
   
   #Copy for Lyndridge Mall
    auto_value_D160 = auto_sheet['D160'].value
    auto_value_A172 = auto_sheet['A172'].value
    integer_part_A172 = int(re.search(r'\d+', auto_value_A172).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_D158= auto_sheet['D158'].value
    auto_value_D182 = auto_sheet['D182'].value
    auto_value_D181 = auto_sheet['D181'].value

    #copy for Melville
    auto_value_D209 = auto_sheet['D209'].value
    auto_value_A221 = auto_sheet['A221'].value
    integer_part_A221 = int(re.search(r'\d+', auto_value_A221).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_M201= auto_sheet['M201'].value
    auto_value_D231 = auto_sheet['D231'].value
    auto_value_D230 = auto_sheet['D230'].value

    #copy for Cambridge
    auto_value_D257 = auto_sheet['D257'].value
    auto_value_A269 = auto_sheet['A269'].value
    integer_part_A269 = int(re.search(r'\d+', auto_value_A269).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_O249= auto_sheet['O249'].value
    auto_value_D279 = auto_sheet['D279'].value
    auto_value_D278 = auto_sheet['D278'].value


    #copy for Queenswood
    auto_value_D305 = auto_sheet['D305'].value
    auto_value_A317 = auto_sheet['A317'].value
    integer_part_A317 = int(re.search(r'\d+', auto_value_A317).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_Q297= auto_sheet['Q297'].value
    auto_value_D327 = auto_sheet['D327'].value
    auto_value_D326 = auto_sheet['D326'].value


    #copy for waterkloof
    auto_value_S355 = auto_sheet['S355'].value
    auto_value_A366 = auto_sheet['A366'].value
    integer_part_A366 = int(re.search(r'\d+', auto_value_A366).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_S346= auto_sheet['S346'].value
    auto_value_D376 = auto_sheet['D376'].value
    auto_value_D375 = auto_sheet['D375'].value

    #copy for Stellenbosch
    auto_value_D402 = auto_sheet['D402'].value
    auto_value_A414 = auto_sheet['A414'].value
    integer_part_A414 = int(re.search(r'\d+', auto_value_A414).group()) #filter the invoice of hartenbros so it copies only the int value on cell A25
    auto_value_U394= auto_sheet['U394'].value
    auto_value_D424 = auto_sheet['D424'].value
    auto_value_D423 = auto_sheet['D423'].value





    # Paste the data from cell D62 into cell B5 in the new sheet
    ws['B5'].value = auto_value_D13
    ws['B5'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J5'].value = integer_part_A25
    ws['J5'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V5'].value = auto_value_B5
    ws['V5'].alignment = Alignment(horizontal='right')
    ws['Y5'].value = auto_value_D35
    ws['Y5'].alignment = Alignment(horizontal='right')
    ws['AB5'].value = auto_value_D34 
    ws['AB5'].alignment = Alignment(horizontal='right')

        # Paste the data from cell D62 into cell B5 in the new sheet
    ws['B6'].value = auto_value_G63
    ws['B6'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J6'].value = integer_part_A74
    ws['J6'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V6'].value = auto_value_G54
    ws['V6'].alignment = Alignment(horizontal='right')
    ws['Y6'].value = auto_value_D84
    ws['Y6'].alignment = Alignment(horizontal='right')
    ws['AB6'].value = auto_value_D83
    ws['AB6'].alignment = Alignment(horizontal='right')


    ws['B7'].value = auto_value_D111
    ws['B7'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J7'].value = integer_part_A123
    ws['J7'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V7'].value = auto_value_I103
    ws['V7'].alignment = Alignment(horizontal='right')
    ws['Y7'].value = auto_value_D133
    ws['Y7'].alignment = Alignment(horizontal='right')
    ws['AB7'].value = auto_value_D132
    ws['AB7'].alignment = Alignment(horizontal='right')


    ws['B9'].value = auto_value_D160
    ws['B9'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J9'].value = integer_part_A172
    ws['J9'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V9'].value = auto_value_D158
    ws['V9'].alignment = Alignment(horizontal='right')
    ws['Y9'].value = auto_value_D182
    ws['Y9'].alignment = Alignment(horizontal='right')
    ws['AB9'].value = auto_value_D181
    ws['AB9'].alignment = Alignment(horizontal='right')

    ws['B11'].value = auto_value_D209
    ws['B11'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J11'].value = integer_part_A221
    ws['J11'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V11'].value = auto_value_M201
    ws['V11'].alignment = Alignment(horizontal='right')
    ws['Y11'].value = auto_value_D231
    ws['Y11'].alignment = Alignment(horizontal='right')
    ws['AB11'].value = auto_value_D230
    ws['AB11'].alignment = Alignment(horizontal='right')

    ws['B12'].value = auto_value_D257
    ws['B12'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J12'].value = integer_part_A269
    ws['J12'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V12'].value = auto_value_O249
    ws['V12'].alignment = Alignment(horizontal='right')
    ws['Y12'].value = auto_value_D279
    ws['Y12'].alignment = Alignment(horizontal='right')
    ws['AB12'].value = auto_value_D278
    ws['AB12'].alignment = Alignment(horizontal='right')


    ws['B13'].value = auto_value_D305
    ws['B13'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J13'].value = integer_part_A317
    ws['J13'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V13'].value = auto_value_Q297
    ws['V13'].alignment = Alignment(horizontal='right')
    ws['Y13'].value = auto_value_D327
    ws['Y13'].alignment = Alignment(horizontal='right')
    ws['AB13'].value = auto_value_D326
    ws['AB13'].alignment = Alignment(horizontal='right')


    ws['B14'].value = auto_value_S355
    ws['B14'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J14'].value = integer_part_A366
    ws['J14'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V14'].value = auto_value_S346
    ws['V14'].alignment = Alignment(horizontal='right')
    ws['Y14'].value = auto_value_D376
    ws['Y14'].alignment = Alignment(horizontal='right')
    ws['AB14'].value = auto_value_D326
    ws['AB14'].alignment = Alignment(horizontal='right')

    ws['B15'].value = auto_value_D402
    ws['B15'].alignment = Alignment(horizontal='right') #align data to the right
    ws['J15'].value = integer_part_A414
    ws['J15'].alignment = Alignment(horizontal='right')#align data to the right
    ws['V15'].value = auto_value_U394
    ws['V15'].alignment = Alignment(horizontal='right')
    ws['Y15'].value = auto_value_D424
    ws['Y15'].alignment = Alignment(horizontal='right')
    ws['AB15'].value = auto_value_D423
    ws['AB15'].alignment = Alignment(horizontal='right')



    auto_wb.close()  # Close the auto.xlsx workbook




    wb.save("PeV Weekly Summary Report 2023.xlsx")

#this section we are converting pdf to excel(extended program)
#the data will be stored on auto.xlsx worksheet
pdf_file = "sample.pdf"
output_excel_file = "auto.xlsx"


#Read PDF and convert to excel
tables =tabula.read_pdf(pdf_file, pages="all", multiple_tables=True)

#combine all tables into a single DataFrame (if multiple tables are present)
df = pd.concat(tables, ignore_index=True)

#save the DataFrame to Excel
df.to_excel(output_excel_file,index=False)

print(f"PDF '{pdf_file}' converted to Excel '{output_excel_file}' successfully")


target_time = time(2,1)  # Set the target time to anytime
wb = load_workbook('PeV Weekly Summary Report 2023.xlsx') #loading a workbook
sheet1 = wb['sheet1']

while True:
    current_time = datetime.now().time()
    
    if current_time >= target_time:
        create_new_sheet(wb)
        print("New sheet created at", current_time)
        break
    
    t.sleep(60)  # Wait for 1 minute before checking again

print(wb)

